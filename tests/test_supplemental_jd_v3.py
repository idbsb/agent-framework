from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.api.app import emerging_job_detail, emerging_jobs
from src.emerging.supplemental_pipeline import parse_supplemental_text


ROOT = Path(__file__).resolve().parents[1]


class SupplementalParserUnitTest(unittest.TestCase):
    def test_supported_heading_number_formats(self):
        text = "\n".join([
            "①岗位甲——企业甲", "岗位职责：开发系统", "任职要求：熟悉Python",
            "⑩①岗位乙——企业乙", "职位描述：完成评测", "职位要求：本科",
            "②⑩岗位丙——企业丙", "工作职责：平台运维", "岗位要求：三年经验",
            "34.岗位丁——企业丁", "职责描述：算法研发", "任职资格：硕士",
            "49.岗位戊——企业戊", "Responsibilities: testing", "Qualifications: Python",
        ])
        records, failures = parse_supplemental_text(text, source_name="fixture.txt")
        self.assertEqual(len(records), 5)
        self.assertEqual(failures, [])
        self.assertEqual(records[0]["raw_job_title"], "岗位甲")
        self.assertEqual(records[-1]["company_raw"], "企业戊")
        self.assertIn("开发系统", records[0]["responsibilities_raw"])
        self.assertIn("Python", records[0]["requirements_raw"])


class SupplementalArtifactsIntegrationTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.records = json.loads((ROOT / "data/external/supplemental_jd_v3.json").read_text(encoding="utf-8"))["records"]
        cls.v2 = json.loads((ROOT / "outputs/emerging_jobs_v2.json").read_text(encoding="utf-8"))

    def test_all_49_records_are_conserved(self):
        self.assertEqual(len(self.records), 49)
        self.assertEqual(sum(not row.get("parse_failed", False) for row in self.records), 49)

    def test_evidence_ids_are_continuous_and_unique(self):
        expected = [f"SUP-JD-{index:03d}" for index in range(1, 50)]
        actual = [row["evidence_id"] for row in self.records]
        self.assertEqual(actual, expected)
        self.assertEqual(len(actual), len(set(actual)))

    def test_required_structure_and_source_integrity(self):
        for row in self.records:
            self.assertTrue(row["raw_job_title"])
            self.assertTrue(row["company_raw"])
            self.assertTrue(row["raw_text"])
            self.assertLessEqual(row["source_start_line"], row["source_end_line"])
            self.assertEqual(row["data_version"], "supplemental_jd_v3")
            self.assertFalse(row["source_url"])
            self.assertEqual(row["source_verification_status"], "待补来源")

    def test_titles_companies_sections_and_skills_are_extracted(self):
        self.assertTrue(all(row["raw_job_title"] and row["company_raw"] for row in self.records))
        first = self.records[0]
        self.assertIn("数据基础设施规划", first["responsibilities_raw"])
        self.assertIn("本科及以上学历", first["requirements_raw"])
        self.assertIn("ROS2", first["bonus_requirements_raw"])
        self.assertIn("Python", first["required_skills"])
        product = self.records[9]
        self.assertIn("专业技能", product["raw_text"])
        self.assertTrue(product["requirements_raw"])
        self.assertTrue(product["bonus_requirements_raw"])

    def test_every_record_has_mapping_and_dedup_conclusion(self):
        valid_mapping = {"exact_evidence", "job_family_support", "new_candidate", "manual_review"}
        valid_duplicate = {"unique", "duplicate", "possible_duplicate"}
        self.assertTrue(all(row["mapping_type"] in valid_mapping for row in self.records))
        self.assertTrue(all(row["duplicate_status"] in valid_duplicate for row in self.records))

    def test_known_duplicate_is_not_counted_twice(self):
        first = self.records[11]
        repeat = self.records[47]
        self.assertEqual(first["raw_job_title"], repeat["raw_job_title"])
        self.assertEqual(repeat["duplicate_status"], "duplicate")
        self.assertEqual(repeat["duplicate_of"], "SUP-JD-012")
        self.assertFalse(repeat["count_in_statistics"])

    def test_old_ids_are_preserved_and_new_ids_are_continuous(self):
        candidates = self.v2["candidates"]
        ids = [row["candidate_id"] for row in candidates]
        self.assertEqual(ids[:11], [f"EMERGING-{index:03d}" for index in range(1, 12)])
        numeric = [int(value.split("-")[-1]) for value in ids]
        self.assertEqual(numeric, list(range(1, max(numeric) + 1)))
        old_v1 = json.loads((ROOT / "outputs/emerging_jobs_v1.json").read_text(encoding="utf-8"))["candidates"]
        self.assertEqual(
            [(row["candidate_id"], row["candidate_name"]) for row in candidates[:11]],
            [(row["candidate_id"], row["candidate_name"]) for row in old_v1],
        )

    def test_old_mapping_and_new_clusters_follow_role_boundaries(self):
        self.assertEqual(self.records[4]["primary_emerging_job_id"], "EMERGING-004")
        self.assertEqual(self.records[7]["primary_emerging_job_id"], "EMERGING-005")
        self.assertEqual(self.records[25]["primary_emerging_job_id"], "EMERGING-011")
        self.assertEqual(self.records[5]["primary_emerging_job_id"], self.records[6]["primary_emerging_job_id"])
        self.assertEqual(len({self.records[index]["primary_emerging_job_id"] for index in (32, 33, 34)}), 1)
        self.assertNotEqual(self.records[35]["primary_emerging_job_id"], self.records[34]["primary_emerging_job_id"])
        self.assertNotEqual(self.records[39]["primary_emerging_job_id"], self.records[36]["primary_emerging_job_id"])

    def test_singleton_new_candidates_remain_weak(self):
        for candidate in self.v2["candidates"][11:]:
            if candidate["counted_evidence_count"] <= 1:
                self.assertEqual(candidate["confidence_v2"], "弱候选/待观察")

    def test_all_frozen_workbooks_are_unchanged(self):
        audit = json.loads((ROOT / "outputs/frozen_file_hash_audit_v3.json").read_text(encoding="utf-8"))
        self.assertTrue(audit["unchanged"])
        self.assertEqual(audit["before"], audit["after"])
        self.assertEqual(audit["file_count"], 10)

    def test_api_list_and_new_detail_are_compatible(self):
        payload = emerging_jobs()
        self.assertEqual(payload["data_version"], "supplemental_jd_v3")
        new_candidate = next(row for row in payload["candidates"] if int(row["candidate_id"].split("-")[-1]) > 11)
        detail = emerging_job_detail(new_candidate["candidate_id"])
        self.assertEqual(detail["candidate_id"], new_candidate["candidate_id"])

    def test_json_and_excel_record_counts_match(self):
        workbook = load_workbook(ROOT / "data/external/supplemental_jd_v3.xlsx", read_only=True, data_only=True)
        try:
            def populated_rows(sheet_name: str) -> int:
                return sum(1 for row in workbook[sheet_name].iter_rows(values_only=True) if any(value not in (None, "") for value in row)) - 1

            self.assertEqual(populated_rows("49条补充JD明细"), len(self.records))
            self.assertEqual(populated_rows("JD原文"), len(self.records))
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()

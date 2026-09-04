from __future__ import annotations

import hashlib
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BATCH_ID = "BATCH-20260904"
GRAPH_VERSION = "Graph V2"
BASELINE_GRAPH_VERSION = "Graph V1"


def _text(value: object) -> str:
    return str(value or "").strip()


def _iso(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()[:10]
    return _text(value)


def _split(value: object) -> list[str]:
    values = re.split(r"[;,，、；|/\n]+", _text(value))
    return list(dict.fromkeys(item.strip() for item in values if len(item.strip()) > 1))


def _ids(value: object) -> list[str]:
    return list(dict.fromkeys(re.findall(r"JD\d{3}", _text(value), re.I)))


class IncrementalDataService:
    """Read-only, schema-tolerant gateway for the six-file incremental batch.

    Files are discovered by sheet signatures rather than exact filenames so a later batch can
    be dropped into a new directory without changing the application code.
    """

    def __init__(self, project_root: Path, skill_index: Any | None = None):
        self.project_root = project_root
        self.batch_dir = project_root / "data" / "incremental" / "batch_20260904"
        self.skill_index = skill_index
        self._cache: dict[str, Any] | None = None

    @staticmethod
    def _read(path: Path) -> dict[str, list[dict[str, Any]]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        result: dict[str, list[dict[str, Any]]] = {}
        try:
            for sheet in workbook.worksheets:
                iterator = sheet.iter_rows(values_only=True)
                headers: list[str] | None = None
                rows: list[dict[str, Any]] = []
                for values in iterator:
                    if not any(value not in (None, "") for value in values):
                        continue
                    if headers is None:
                        headers = [_text(value) for value in values]
                        continue
                    row = {headers[i]: value for i, value in enumerate(values) if i < len(headers) and headers[i]}
                    if any(value not in (None, "") for value in row.values()):
                        rows.append(row)
                result[sheet.title] = rows
        finally:
            workbook.close()
        return result

    def load(self) -> dict[str, Any]:
        if self._cache is not None:
            return self._cache
        workbooks: list[dict[str, Any]] = []
        sheets: dict[str, list[dict[str, Any]]] = {}
        for path in sorted(self.batch_dir.glob("*.xlsx")):
            book = self._read(path)
            digest = hashlib.sha256(path.read_bytes()).hexdigest().upper()
            workbooks.append({
                "file_name": path.name,
                "sha256": digest,
                "sheets": [{"sheet_name": name, "row_count": len(rows), "fields": list(rows[0]) if rows else []} for name, rows in book.items()],
            })
            sheets.update(book)
        expected = {"JD主数据", "来源审计", "时间窗口统计", "岗位时间对比", "技能变化明细", "JD证据映射", "新兴岗位趋势", "政策文件", "行业报告", "多源交叉验证", "检索日志"}
        missing = sorted(expected - set(sheets))
        if len(workbooks) != 6 or missing:
            raise RuntimeError(f"增量批次不完整：文件 {len(workbooks)}/6，缺少工作表 {missing}")
        self._cache = {"workbooks": workbooks, "sheets": sheets}
        return self._cache

    def raw_jds(self) -> list[dict[str, Any]]:
        return self.load()["sheets"]["JD主数据"]

    def counted_jds(self) -> list[dict[str, Any]]:
        rows = self.raw_jds()
        return [row for row in rows if not _text(row.get("canonical_jd_id")) or _text(row.get("canonical_jd_id")) == _text(row.get("JD编号"))]

    def standardized_jds(self) -> list[dict[str, Any]]:
        result = []
        for row in self.counted_jds():
            raw_id = _text(row.get("JD编号"))
            result.append({
                "jd_id": f"{BATCH_ID}-{raw_id}", "evidence_id": raw_id,
                "original_job_title": _text(row.get("原始岗位名称")), "standard_job_title": _text(row.get("岗位标准分类")),
                "job_family": _text(row.get("岗位分组")), "technical_domain": _text(row.get("所属行业")),
                "company": _text(row.get("企业名称")), "city": _text(row.get("工作地点")),
                "responsibilities": _text(row.get("岗位职责")), "required_skills_raw": _text(row.get("任职要求")) + "; " + _text(row.get("技能原文")),
                "bonus_skills_raw": _text(row.get("加分项/优先条件")), "education": _text(row.get("学历要求")),
                "experience": _text(row.get("经验要求")), "original_experience": _text(row.get("经验要求")),
                "source": _text(row.get("来源平台")), "source_url": _text(row.get("来源URL")),
                "publish_date": _iso(row.get("发布日期")), "standardization_status": _text(row.get("真实性状态")),
                "original_row_number": raw_id, "raw_text": "\n".join(_text(row.get(key)) for key in ("岗位职责", "任职要求", "加分项/优先条件")),
                "batch_id": BATCH_ID, "data_protection": "INCREMENTAL_READ_ONLY",
            })
        return result

    def changes(self) -> list[dict[str, Any]]:
        skill_rows = self.load()["sheets"]["技能变化明细"]
        job_rows = self.load()["sheets"]["岗位时间对比"]
        status_by_job: dict[str, list[str]] = defaultdict(list)
        support_by_job: dict[str, set[str]] = defaultdict(set)
        for row in job_rows:
            job = _text(row.get("岗位")); status_by_job[job].append(_text(row.get("样本状态"))); support_by_job[job].update(_ids(row.get("支撑JD")))
        grouped: dict[str, dict[str, Any]] = {}
        for row in skill_rows:
            job, skill, direction = _text(row.get("岗位")), _text(row.get("技能")), _text(row.get("变化方向"))
            item = grouped.setdefault(job, {"job_title": job, "new_skills": [], "growing_skills": [], "stable_skills": [], "declining_skills": [], "clues": [], "support_jd_ids": set()})
            item["support_jd_ids"].update(_ids(row.get("支撑JD")))
            if "新增" in direction: item["new_skills"].append(skill)
            elif "增强" in direction or "上升" in direction: item["growing_skills"].append(skill)
            elif "稳定" in direction: item["stable_skills"].append(skill)
            elif "下降" in direction: item["declining_skills"].append(skill)
            else: item["clues"].append(skill)
        result = []
        for job, item in grouped.items():
            insufficient = any("不足" in status or "insufficient" in status.lower() for status in status_by_job[job])
            item["support_jd_ids"].update(support_by_job[job]); item["support_jd_ids"] = sorted(item["support_jd_ids"])
            item.update(sample_insufficient=insufficient, sample_status="样本不足，当前仅作为变化线索，不形成显著趋势判断。" if insufficient else "可比较", batch_id=BATCH_ID)
            result.append(item)
        return result

    def emerging(self) -> list[dict[str, Any]]:
        result = []
        for index, row in enumerate(self.load()["sheets"]["新兴岗位趋势"], 1):
            result.append({
                "candidate_id": f"INC-EMERGING-{index:03d}", "candidate_name": _text(row.get("岗位候选")),
                "status": "新兴岗位候选", "jd_count": int(row.get("JD数量") or 0), "company_count": int(row.get("企业数量") or 0),
                "source_count": int(row.get("来源数量") or 0), "first_seen": _iso(row.get("首次观察时间")), "last_seen": _iso(row.get("最近观察时间")),
                "core_skills": _split(row.get("核心技能")), "distinguishing_skills": _split(row.get("新增/差异技能")),
                "similar_existing_jobs": _split(row.get("相关既有岗位")), "notice": _text(row.get("备注")), "batch_id": BATCH_ID,
            })
        return result

    def external_evidence(self) -> list[dict[str, Any]]:
        result = []
        for source_type, sheet_name in (("政策证据", "政策文件"), ("行业报告证据", "行业报告")):
            for row in self.load()["sheets"][sheet_name]:
                result.append({"evidence_id": _text(row.get("编号")), "source_type": source_type,
                    "title": _text(row.get("文件名称") or row.get("报告名称")), "publisher": _text(row.get("发布机构")),
                    "published_date": _iso(row.get("发布日期")), "summary": _text(row.get("与项目相关的原文摘要") or row.get("关键结论")),
                    "supports": _text(row.get("支持的岗位/技能方向") or row.get("与项目相关的技能/岗位趋势")), "url": _text(row.get("原始URL"))})
        return result

    def cross_validation(self) -> list[dict[str, Any]]:
        return [{"trend": _text(r.get("岗位/技能趋势")), "jd_evidence_ids": _ids(r.get("JD证据")),
                 "policy_evidence_ids": re.findall(r"P\d{3}", _text(r.get("政策证据")), re.I),
                 "report_evidence_ids": re.findall(r"R\d{3}", _text(r.get("行业报告证据")), re.I),
                 "other_evidence": _text(r.get("GitHub技术证据")), "summary": _text(r.get("综合说明")), "strength": _text(r.get("证据强度"))}
                for r in self.load()["sheets"]["多源交叉验证"]]

    def audit(self) -> dict[str, Any]:
        statuses = Counter(_text(r.get("真实性状态")) or "UNKNOWN" for r in self.raw_jds())
        logs = self.load()["sheets"]["检索日志"]
        inclusion = Counter("已纳入" if _text(r.get("是否纳入")) == "是" else "未纳入" for r in logs)
        excluded = Counter(_text(r.get("未纳入原因")) for r in logs if _text(r.get("是否纳入")) != "是")
        return {"status_counts": dict(statuses), "inclusion_counts": dict(inclusion), "exclusion_reasons": dict(excluded), "audit_record_count": len(logs)}

    def skill_names(self) -> dict[str, set[str]]:
        by_job: dict[str, set[str]] = defaultdict(set)
        for row in self.counted_jds():
            by_job[_text(row.get("岗位标准分类"))].update(_split(row.get("技能原文")))
        return by_job

    def overview(self, baseline_jd_count: int, baseline_jobs: set[str], baseline_graph: dict[str, int]) -> dict[str, Any]:
        counted = self.counted_jds(); all_rows = self.raw_jds(); skill_names = self.skill_names()
        new_jobs = set(skill_names) - baseline_jobs
        companies = {_text(r.get("企业名称")) for r in counted if _text(r.get("企业名称")) and "无法确认" not in _text(r.get("企业名称"))}
        sources = {_text(r.get("来源平台")) for r in counted if _text(r.get("来源平台"))}
        dated = sum(bool(_iso(r.get("发布日期"))) for r in counted)
        new_skills = set().union(*skill_names.values()) if skill_names else set()
        return {"batch_id": BATCH_ID, "graph_version": GRAPH_VERSION, "baseline_graph_version": BASELINE_GRAPH_VERSION,
            "updated_at": "2026-09-04", "files_loaded": len(self.load()["workbooks"]), "workbooks": self.load()["workbooks"],
            "baseline": {"jd_count": baseline_jd_count, **baseline_graph},
            "incremental": {"raw_jd_count": len(all_rows), "jd_count": len(counted), "duplicate_excluded_count": len(all_rows)-len(counted),
                "source_count": len(sources), "company_count": len(companies), "job_count": len(skill_names), "dated_jd_count": dated,
                "dated_jd_ratio": round(dated / len(counted), 4) if counted else 0, "new_job_count": len(new_jobs), "observed_skill_count": len(new_skills),
                "changed_job_count": len(self.changes()), "emerging_candidate_count": len(self.emerging())},
            "current": {"jd_count": baseline_jd_count + len(counted), "job_count": len(baseline_jobs | set(skill_names))},
            "recruitment_sources": dict(Counter(_text(r.get("来源平台")) for r in counted)), "audit": self.audit(),
            "external_source_counts": {"政策文件": len(self.load()["sheets"]["政策文件"]), "行业报告": len(self.load()["sheets"]["行业报告"]), "交叉验证": len(self.cross_validation())}}

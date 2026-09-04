from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import yaml
from openpyxl import load_workbook


class DataConfigurationError(RuntimeError):
    pass


def _scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return "" if value is None else value


class DataLoader:
    """Single configurable read-only gateway for all formal datasets."""

    FROZEN_KEYS = (
        "standardized_jd_dataset",
        "standard_job_title_mapping",
        "standard_skill_dictionary",
        "standardized_resume_testset",
    )

    def __init__(self, config_path: str | Path | None = None):
        default = Path(__file__).resolve().parents[1] / "config" / "data_sources.yaml"
        self.config_path = Path(config_path or default).resolve()
        self.project_root = self.config_path.parents[1]
        self.sources = self._read_yaml(self.config_path)
        self.field_mapping = self._read_yaml(self.project_root / "config" / "field_mapping.yaml")
        self.version = self._read_yaml(self.project_root / "config" / "version.yaml")

    @staticmethod
    def _read_yaml(path: Path) -> dict[str, Any]:
        if not path.exists():
            raise DataConfigurationError(f"配置文件不存在：{path}")
        value = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        if not isinstance(value, dict):
            raise DataConfigurationError(f"配置文件顶层必须为对象：{path}")
        return value

    def resolve_path(self, source_key: str) -> Path:
        source = self.sources.get(source_key)
        if not isinstance(source, dict) or not source.get("path"):
            raise DataConfigurationError(f"数据源未配置 path：{source_key}")
        path = Path(source["path"])
        return path.resolve() if path.is_absolute() else (self.project_root / path).resolve()

    def output_dir(self) -> Path:
        path = Path(self.sources.get("outputs_dir", "outputs"))
        return path.resolve() if path.is_absolute() else (self.project_root / path).resolve()

    def reports_dir(self) -> Path:
        path = Path(self.sources.get("reports_dir", "reports"))
        return path.resolve() if path.is_absolute() else (self.project_root / path).resolve()

    @staticmethod
    def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
        if not path.exists():
            raise DataConfigurationError(f"数据文件不存在：{path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise DataConfigurationError(f"{path.name} 中不存在工作表：{sheet_name}")
            sheet = workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = [str(x).strip() if x is not None else "" for x in next(iterator)]
            except StopIteration:
                return []
            rows = []
            for excel_row, values in enumerate(iterator, start=2):
                row = {headers[i]: _scalar(value) for i, value in enumerate(values) if i < len(headers) and headers[i]}
                if any(value not in (None, "") for value in row.values()):
                    row["_excel_row"] = excel_row
                    rows.append(row)
            return rows
        finally:
            workbook.close()

    def _load_mapped(self, source_key: str) -> list[dict[str, Any]]:
        source = self.sources[source_key]
        raw = self.read_sheet(self.resolve_path(source_key), source["sheet"])
        mapping = self.field_mapping.get(source_key, {})
        if not mapping:
            raise DataConfigurationError(f"字段映射不存在：{source_key}")
        missing = [column for column in mapping.values() if raw and column not in raw[0]]
        if missing:
            raise DataConfigurationError(f"{source_key} 缺少字段：{missing}")
        return [
            {**{internal: row.get(column, "") for internal, column in mapping.items()}, "_excel_row": row["_excel_row"]}
            for row in raw
        ]

    def load_jds(self) -> list[dict[str, Any]]:
        return self._load_mapped("standardized_jd_dataset")

    def load_job_analysis_jds(self) -> list[dict[str, Any]]:
        """Return the immutable formal corpus plus approved supplemental statistics rows.

        The supplemental source and its existing candidate mapping remain untouched. Confirmed
        duplicates stay in the protected source for auditability but are not counted twice.
        """
        rows = self.load_jds()
        source = self.sources.get("supplemental_job_analysis")
        if not isinstance(source, dict):
            return rows
        source_path = Path(str(source.get("path", "")))
        mapping_path = Path(str(source.get("candidate_mapping_path", "")))
        source_path = source_path if source_path.is_absolute() else self.project_root / source_path
        mapping_path = mapping_path if mapping_path.is_absolute() else self.project_root / mapping_path
        if not source_path.exists() or not mapping_path.exists():
            raise DataConfigurationError("岗位分析补充JD或其既有候选映射不存在")
        payload = json.loads(source_path.read_text(encoding="utf-8"))
        mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
        records = payload.get("records")
        candidates = mapping.get("candidates")
        if not isinstance(records, list) or not isinstance(candidates, list):
            raise DataConfigurationError("岗位分析补充JD或候选映射结构无效")
        candidate_names = {
            str(item.get("candidate_id", "")).strip(): str(item.get("candidate_name", "")).strip()
            for item in candidates if isinstance(item, dict)
        }
        include_flag = str(source.get("include_flag") or "count_in_statistics")
        protection_label = str(source.get("protection_label") or "PROTECTED_NEW_DATA")
        supplemental = []
        for record in records:
            if not isinstance(record, dict) or not record.get(include_flag):
                continue
            candidate_id = str(record.get("primary_emerging_job_id", "")).strip()
            title = candidate_names.get(candidate_id, "")
            if not title:
                raise DataConfigurationError(f"补充JD缺少既有岗位候选映射：{record.get('evidence_id', '')}")
            supplemental.append({
                "jd_id": str(record.get("evidence_id", "")).strip(),
                "original_job_title": str(record.get("raw_job_title", "")).strip(),
                "standard_job_title": title,
                "job_family": str(record.get("job_family", "")).strip(),
                "technical_domain": str(record.get("technical_domain", "")).strip(),
                "company": str(record.get("company_normalized", "")).strip(),
                "city": str(record.get("location", "")).strip(),
                "responsibilities": str(record.get("responsibilities_raw", "")).strip(),
                "required_skills_raw": str(record.get("requirements_raw", "")).strip(),
                "bonus_skills_raw": str(record.get("bonus_requirements_raw", "")).strip(),
                "education": str(record.get("education_requirement", "")).strip(),
                "experience": str(record.get("experience_requirement", "")).strip(),
                "original_experience": str(record.get("experience_requirement", "")).strip(),
                "source": str(record.get("source_name", "")).strip(),
                "source_url": str(record.get("source_url", "")).strip(),
                "standardization_status": str(record.get("mapping_type", "")).strip(),
                "original_row_number": record.get("source_start_line", ""),
                "raw_text": str(record.get("raw_text", "")).strip(),
                "candidate_id": candidate_id,
                "data_protection": protection_label,
            })
        from .integration.incremental_data import IncrementalDataService
        incremental = IncrementalDataService(self.project_root).standardized_jds()
        ids = [row.get("jd_id") for row in rows + supplemental + incremental]
        if len(ids) != len(set(ids)):
            raise DataConfigurationError("岗位分析组合输入存在重复JD编号")
        return rows + supplemental + incremental

    def load_incremental_jds(self) -> list[dict[str, Any]]:
        from .integration.incremental_data import IncrementalDataService
        return IncrementalDataService(self.project_root).standardized_jds()

    def job_analysis_data_version(self) -> str:
        base = str(self.version.get("data_version") or "unknown")
        source = self.sources.get("supplemental_job_analysis")
        if not isinstance(source, dict):
            return base
        path = Path(str(source.get("path", "")))
        path = path if path.is_absolute() else self.project_root / path
        payload = json.loads(path.read_text(encoding="utf-8"))
        supplemental = str(payload.get("data_version") or "supplemental")
        return f"{base}+{supplemental}+batch_20260904"

    def load_job_title_mapping(self) -> list[dict[str, Any]]:
        return self._load_mapped("standard_job_title_mapping")

    def load_resumes(self) -> list[dict[str, Any]]:
        return self._load_mapped("standardized_resume_testset")

    def load_skill_dictionary(self) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        source = self.sources["standard_skill_dictionary"]
        path = self.resolve_path("standard_skill_dictionary")
        skills = self.read_sheet(path, source["standard_skills_sheet"])
        aliases = self.read_sheet(path, source["aliases_sheet"])
        return skills, aliases

    def load_job_skill_gold(self) -> list[dict[str, Any]]:
        source = self.sources.get("job_skill_gold")
        if not source:
            return []
        return self.read_sheet(self.resolve_path("job_skill_gold"), source["sheet"])

    def load_runtime_skill_dictionary(self) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        """Add reviewed runtime entries without changing the frozen source data."""
        skills, aliases = self.load_skill_dictionary()
        path = self.project_root / "config" / "skill_dictionary_extensions.json"
        extension = json.loads(path.read_text(encoding="utf-8"))
        ids = {row["skill_id"] for row in skills}
        names = {row["标准技能名称"].casefold() for row in skills}
        for row in extension["skills"]:
            if row["skill_id"] in ids or row["标准技能名称"].casefold() in names:
                raise DataConfigurationError("增量技能 ID 或名称与现有词典冲突，请人工合并")
            skills.append(row)
            ids.add(row["skill_id"])
            names.add(row["标准技能名称"].casefold())
        for row in extension["aliases"]:
            if row["skill_id"] not in ids:
                raise DataConfigurationError("增量别名指向不存在的技能 ID")
            aliases.append(row)
        return skills, aliases

    def source_locations(self) -> dict[str, str]:
        return {key: str(self.resolve_path(key)) for key in self.FROZEN_KEYS}

    @staticmethod
    def sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest().upper()

    def frozen_hashes(self) -> dict[str, str]:
        return {key: self.sha256(self.resolve_path(key)) for key in self.FROZEN_KEYS}

    @staticmethod
    def duplicate_values(rows: Iterable[dict[str, Any]], key: str) -> list[str]:
        seen, duplicates = set(), set()
        for row in rows:
            value = str(row.get(key, "")).strip()
            if value in seen:
                duplicates.add(value)
            seen.add(value)
        return sorted(duplicates)

    def validate_frozen(self) -> dict[str, Any]:
        jds = self.load_jds()
        resumes = self.load_resumes()
        skills, aliases = self.load_skill_dictionary()
        skill_ids = {str(row.get("skill_id", "")).strip() for row in skills}
        alias_unknown = sorted({str(row.get("skill_id", "")).strip() for row in aliases} - skill_ids - {""})
        result = {
            "jd_duplicate_ids": self.duplicate_values(jds, "jd_id"),
            "resume_duplicate_ids": self.duplicate_values(resumes, "resume_id"),
            "skill_duplicate_ids": self.duplicate_values(skills, "skill_id"),
            "standard_skill_name_duplicates": self.duplicate_values(skills, "标准技能名称"),
            "alias_unknown_skill_ids": alias_unknown,
            "row_counts": {"jds": len(jds), "resumes": len(resumes), "skills": len(skills), "aliases": len(aliases)},
            "source_locations": self.source_locations(),
            "data_version": self.version.get("data_version"),
            "schema_version": self.version.get("schema_version"),
        }
        result["passed"] = not any(result[key] for key in (
            "jd_duplicate_ids", "resume_duplicate_ids", "skill_duplicate_ids",
            "standard_skill_name_duplicates", "alias_unknown_skill_ids",
        ))
        return result


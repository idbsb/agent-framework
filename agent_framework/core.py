from __future__ import annotations

import json
import re
import sqlite3
from io import BytesIO
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable


NOW = lambda: datetime.now(timezone.utc).isoformat(timespec="seconds")


class Database:
    def __init__(self, path: str | Path = "data/challenge_cup.db"):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(self.path)
        self.conn.row_factory = sqlite3.Row

    def init(self) -> None:
        self.conn.executescript("""
        PRAGMA foreign_keys = ON;
        CREATE TABLE IF NOT EXISTS jobs (
          job_id TEXT PRIMARY KEY, original_title TEXT NOT NULL, standard_title TEXT,
          job_cluster TEXT, company TEXT, published_at TEXT, responsibilities TEXT,
          requirements TEXT, education_experience TEXT, source_url TEXT, source_name TEXT,
          status TEXT, raw_json TEXT NOT NULL, imported_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS skills (
          skill_id INTEGER PRIMARY KEY AUTOINCREMENT, canonical_name TEXT UNIQUE NOT NULL,
          category TEXT DEFAULT 'technical', description TEXT
        );
        CREATE TABLE IF NOT EXISTS job_skills (
          job_id TEXT NOT NULL, skill_id INTEGER NOT NULL, frequency INTEGER NOT NULL DEFAULT 1,
          weight REAL NOT NULL, confidence REAL NOT NULL, evidence TEXT NOT NULL,
          extractor TEXT NOT NULL, reviewed_state TEXT DEFAULT 'pending', updated_at TEXT NOT NULL,
          PRIMARY KEY (job_id, skill_id), FOREIGN KEY(job_id) REFERENCES jobs(job_id),
          FOREIGN KEY(skill_id) REFERENCES skills(skill_id)
        );
        CREATE TABLE IF NOT EXISTS job_profiles (
          profile_id INTEGER PRIMARY KEY AUTOINCREMENT, job_cluster TEXT NOT NULL,
          version INTEGER NOT NULL, profile_json TEXT NOT NULL, created_at TEXT NOT NULL,
          UNIQUE(job_cluster, version)
        );
        CREATE TABLE IF NOT EXISTS resumes (
          resume_id TEXT PRIMARY KEY, content TEXT NOT NULL, imported_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS assessments (
          assessment_id INTEGER PRIMARY KEY AUTOINCREMENT, resume_id TEXT NOT NULL, profile_id INTEGER NOT NULL,
          score REAL NOT NULL, matched_json TEXT NOT NULL, missing_json TEXT NOT NULL,
          advice_json TEXT NOT NULL, created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS reviews (
          review_id INTEGER PRIMARY KEY AUTOINCREMENT, job_id TEXT NOT NULL, skill_name TEXT NOT NULL,
          decision TEXT NOT NULL CHECK(decision IN ('confirm','reject','add')), reviewer TEXT,
          reason TEXT, created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS audit_events (
          event_id INTEGER PRIMARY KEY AUTOINCREMENT, event_type TEXT NOT NULL, payload_json TEXT NOT NULL,
          created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS system_state (
          state_key TEXT PRIMARY KEY, value_json TEXT NOT NULL, updated_at TEXT NOT NULL
        );
        """)
        columns = {row[1] for row in self.conn.execute("PRAGMA table_info(jobs)")}
        if "technical_domain" not in columns:
            self.conn.execute("ALTER TABLE jobs ADD COLUMN technical_domain TEXT")
        self.conn.commit()

    def event(self, event_type: str, payload: dict) -> None:
        self.conn.execute("INSERT INTO audit_events(event_type,payload_json,created_at) VALUES(?,?,?)",
                          (event_type, json.dumps(payload, ensure_ascii=False), NOW()))
        self.conn.commit()

    def set_state(self, key: str, value: dict) -> None:
        self.conn.execute("""INSERT INTO system_state(state_key,value_json,updated_at) VALUES(?,?,?)
          ON CONFLICT(state_key) DO UPDATE SET value_json=excluded.value_json,updated_at=excluded.updated_at""",
          (key, json.dumps(value, ensure_ascii=False), NOW()))
        self.conn.commit()

    def get_state(self, key: str, default=None):
        row = self.conn.execute("SELECT value_json FROM system_state WHERE state_key=?", (key,)).fetchone()
        return json.loads(row[0]) if row else default

    def close(self) -> None:
        self.conn.close()


@dataclass
class ExtractedSkill:
    name: str
    evidence: str
    confidence: float = 0.75


class RuleSkillExtractor:
    """Deterministic, explainable extractor; replaceable by an LLM extractor later."""
    name = "rule-taxonomy-v1"

    def __init__(self, taxonomy_path: str | Path | None = None):
        path = Path(taxonomy_path or Path(__file__).with_name("taxonomy.json"))
        self.taxonomy = json.loads(path.read_text(encoding="utf-8"))

    def extract(self, text: str) -> list[ExtractedSkill]:
        normalized = text.casefold()
        found = []
        for canonical, aliases in self.taxonomy.items():
            hits = [a for a in aliases if a.casefold() in normalized]
            if hits:
                found.append(ExtractedSkill(canonical, "；".join(hits), min(0.98, 0.72 + .08 * len(hits))))
        return found


class AgentPipeline:
    def __init__(self, db: Database, extractor: RuleSkillExtractor | None = None):
        self.db, self.extractor = db, extractor or RuleSkillExtractor()

    def import_jobs(self, records: Iterable[dict], source: str = "json") -> dict:
        summary = {"processed": 0, "added": 0, "updated": 0, "unchanged": 0,
                   "added_job_ids": [], "updated_job_ids": [], "source": source, "imported_at": NOW()}
        touched_clusters = set()
        for row in records:
            job_id = str(row.get("job_id") or row.get("id") or f"JD-{summary['processed']+1:04d}")
            title = row.get("original_title") or row.get("原始岗位名") or row.get("职位名称") or "未命名岗位"
            cluster = row.get("job_cluster") or row.get("岗位簇") or row.get("标准岗位名称") or title
            responsibilities = row.get("responsibilities") or row.get("职责摘要") or row.get("岗位职责") or ""
            requirements = row.get("requirements") or row.get("技能摘要") or row.get("任职要求") or ""
            values = {
                "original_title": title,
                "standard_title": row.get("standard_title") or row.get("标准岗位名称") or cluster,
                "job_cluster": cluster,
                "company": row.get("company") or row.get("企业") or row.get("企业名称"),
                "published_at": row.get("published_at") or row.get("发布时间"),
                "responsibilities": responsibilities,
                "requirements": requirements,
                "education_experience": row.get("education_experience") or row.get("学历经验"),
                "source_url": row.get("source_url") or row.get("url") or row.get("招聘来源原始链接"),
                "source_name": row.get("source_name") or row.get("来源"),
                "status": row.get("status") or row.get("状态"),
                "technical_domain": row.get("technical_domain") or row.get("技术领域"),
            }
            # Excel may read a numeric-looking company name (for example "360")
            # as an integer, while SQLite TEXT returns it as a string later.
            # Normalize fields so an identical re-import remains unchanged.
            values = {
                key: None if value is None else str(value)
                for key, value in values.items()
            }
            existing = self.db.conn.execute("""SELECT original_title,standard_title,job_cluster,company,published_at,
              responsibilities,requirements,education_experience,source_url,source_name,status,technical_domain
              FROM jobs WHERE job_id=?""", (job_id,)).fetchone()
            changed = existing is None or any((existing[key] or "") != (values[key] or "") for key in values)
            summary["processed"] += 1
            if existing is None:
                summary["added"] += 1; summary["added_job_ids"].append(job_id)
            elif changed:
                summary["updated"] += 1; summary["updated_job_ids"].append(job_id)
                touched_clusters.add(existing["job_cluster"])
            else:
                summary["unchanged"] += 1
                continue
            self.db.conn.execute("""INSERT INTO jobs(job_id,original_title,standard_title,job_cluster,company,published_at,responsibilities,requirements,education_experience,source_url,source_name,status,raw_json,imported_at,technical_domain) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
              ON CONFLICT(job_id) DO UPDATE SET original_title=excluded.original_title,standard_title=excluded.standard_title,
              job_cluster=excluded.job_cluster,company=excluded.company,published_at=excluded.published_at,
              responsibilities=excluded.responsibilities,requirements=excluded.requirements,education_experience=excluded.education_experience,
              source_url=excluded.source_url,source_name=excluded.source_name,status=excluded.status,raw_json=excluded.raw_json,imported_at=excluded.imported_at,technical_domain=excluded.technical_domain""",
              (job_id, values["original_title"], values["standard_title"], values["job_cluster"], values["company"],
               values["published_at"], values["responsibilities"], values["requirements"], values["education_experience"],
               values["source_url"], values["source_name"], values["status"], json.dumps(row, ensure_ascii=False), NOW(),
               values["technical_domain"]))
            if existing is not None:
                self.db.conn.execute("DELETE FROM job_skills WHERE job_id=? AND reviewed_state='pending'", (job_id,))
            self._extract_job(job_id, responsibilities + "\n" + requirements)
            touched_clusters.add(cluster)
        for cluster_name in touched_clusters:
            if cluster_name:
                self.db.conn.execute("DELETE FROM job_profiles WHERE job_cluster=?", (cluster_name,))
        self.db.conn.commit()
        summary["extractor"] = self.extractor.name
        self.db.event("jobs_imported", summary)
        self.db.set_state("last_import", summary)
        return summary

    def import_excel(self, jd_workbook: bytes, standard_workbook: bytes) -> dict:
        """Import the two agreed first-group Excel sheets without altering either source file."""
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel 导入组件不可用") from exc

        def sheet_rows(raw: bytes):
            ws = load_workbook(BytesIO(raw), read_only=True, data_only=True).active
            rows = ws.iter_rows(values_only=True)
            headers = [str(x).strip() if x is not None else "" for x in next(rows)]
            return [{headers[i]: value for i, value in enumerate(row) if i < len(headers)} for row in rows]

        standards = sheet_rows(standard_workbook)
        exact_by_title, alias_by_title = {}, {}
        for row in standards:
            standard = row.get("标准岗位名称")
            if not standard:
                continue
            for name in (row.get("原始岗位名称"), row.get("清洗后岗位名称"), standard):
                if name:
                    exact_by_title[str(name).strip().casefold()] = row
            for name in str(row.get("岗位别名") or "").replace("、", "，").split("，"):
                if name:
                    alias_by_title.setdefault(str(name).strip().casefold(), row)

        records, unmatched = [], []
        for row in sheet_rows(jd_workbook):
            job_id = row.get("JD编号")
            title = str(row.get("原始岗位名称") or "").strip()
            if not job_id or not title:
                continue
            mapped = exact_by_title.get(title.casefold()) or alias_by_title.get(title.casefold())
            if not mapped:
                unmatched.append({"JD编号": str(job_id), "原始岗位名称": title})
            records.append({
                "id": str(job_id), "原始岗位名": title,
                "标准岗位名称": mapped.get("标准岗位名称") if mapped else title,
                "岗位簇": mapped.get("岗位族") if mapped else "待人工归类",
                "技术领域": mapped.get("技术领域") if mapped else "待人工确认",
                "企业": row.get("企业名称"), "发布时间": str(row.get("发布时间") or ""),
                "职责摘要": row.get("工作职责") or "", "技能摘要": "\n".join(str(row.get(k) or "") for k in ("必备技能", "加分技能")),
                "学历经验": "；".join(str(row.get(k) or "") for k in ("学历要求", "工作经验")),
                "来源": row.get("招聘来源"), "url": row.get("原始链接"), "采集时间": str(row.get("采集时间") or ""), "状态": "Excel导入"
            })
        import_summary = self.import_jobs(records, source="excel")
        result = {**import_summary, "imported": import_summary["processed"], "standard_rows": len(standards),
                  "unmatched": unmatched, "unmatched_count": len(unmatched)}
        self.db.event("excel_imported", result)
        self.db.set_state("last_import", result)
        return result

    def status(self) -> dict:
        row = self.db.conn.execute("""SELECT COUNT(*) job_count,
          COUNT(DISTINCT CASE WHEN job_cluster IS NOT NULL AND job_cluster != '' THEN job_cluster END) cluster_count,
          COUNT(DISTINCT CASE WHEN technical_domain IS NOT NULL AND technical_domain != '' THEN technical_domain END) domain_count
          FROM jobs""").fetchone()
        skill_count = self.db.conn.execute("SELECT COUNT(DISTINCT skill_id) FROM job_skills WHERE reviewed_state != 'rejected'").fetchone()[0]
        return {"job_count": row[0], "cluster_count": row[1], "domain_count": row[2], "skill_count": skill_count,
                "last_import": self.db.get_state("last_import", {}), "last_result": self.db.get_state("last_ui_result", {})}

    def _skill_id(self, name: str) -> int:
        self.db.conn.execute("INSERT OR IGNORE INTO skills(canonical_name) VALUES(?)", (name,))
        return self.db.conn.execute("SELECT skill_id FROM skills WHERE canonical_name=?", (name,)).fetchone()[0]

    def _extract_job(self, job_id: str, text: str) -> None:
        for item in self.extractor.extract(text):
            sid = self._skill_id(item.name)
            prior = self.db.conn.execute("SELECT reviewed_state FROM job_skills WHERE job_id=? AND skill_id=?", (job_id, sid)).fetchone()
            state = prior[0] if prior else "pending"
            confidence = 0.95 if state in ("confirmed", "added") else (0.10 if state == "rejected" else item.confidence)
            self.db.conn.execute("""INSERT INTO job_skills VALUES(?,?,?,?,?,?,?,?,?)
              ON CONFLICT(job_id,skill_id) DO UPDATE SET frequency=excluded.frequency,weight=excluded.weight,
              confidence=excluded.confidence,evidence=excluded.evidence,extractor=excluded.extractor,updated_at=excluded.updated_at""",
              (job_id, sid, 1, 1.0, confidence, item.evidence, self.extractor.name, state, NOW()))

    def review(self, job_id: str, skill_name: str, decision: str, reviewer: str = "", reason: str = "") -> None:
        if decision not in {"confirm", "reject", "add"}: raise ValueError("decision must be confirm/reject/add")
        sid = self._skill_id(skill_name)
        state = {"confirm": "confirmed", "reject": "rejected", "add": "added"}[decision]
        self.db.conn.execute("""INSERT INTO job_skills(job_id,skill_id,frequency,weight,confidence,evidence,extractor,reviewed_state,updated_at)
          VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(job_id,skill_id) DO UPDATE SET reviewed_state=excluded.reviewed_state,
          confidence=excluded.confidence,evidence=excluded.evidence,updated_at=excluded.updated_at""",
          (job_id, sid, 1, 1.0, 0.95 if decision != "reject" else 0.10, "人工复核：" + (reason or decision), "human-review", state, NOW()))
        self.db.conn.execute("INSERT INTO reviews(job_id,skill_name,decision,reviewer,reason,created_at) VALUES(?,?,?,?,?,?)",
                             (job_id, skill_name, decision, reviewer, reason, NOW()))
        self.db.conn.commit(); self.db.event("skill_reviewed", {"job_id": job_id, "skill": skill_name, "decision": decision})

    def build_profile(self, cluster: str) -> dict:
        rows = self.db.conn.execute("""SELECT s.canonical_name, SUM(js.frequency * js.weight * js.confidence) score,
          COUNT(*) jd_count FROM job_skills js JOIN skills s ON s.skill_id=js.skill_id JOIN jobs j ON j.job_id=js.job_id
          WHERE j.job_cluster=? AND js.reviewed_state != 'rejected' GROUP BY s.skill_id ORDER BY score DESC, s.canonical_name""", (cluster,)).fetchall()
        total = self.db.conn.execute("SELECT COUNT(*) FROM jobs WHERE job_cluster=?", (cluster,)).fetchone()[0]
        skills = [{"skill": r[0], "weight": round(r[1] / max(total, 1), 3), "jd_count": r[2]} for r in rows]
        profile = {"job_cluster": cluster, "jd_count": total, "skills": skills, "generated_at": NOW(), "method": "frequency×weight×confidence"}
        version = self.db.conn.execute("SELECT COALESCE(MAX(version),0)+1 FROM job_profiles WHERE job_cluster=?", (cluster,)).fetchone()[0]
        cur = self.db.conn.execute("INSERT INTO job_profiles(job_cluster,version,profile_json,created_at) VALUES(?,?,?,?)",
                                   (cluster, version, json.dumps(profile, ensure_ascii=False), NOW()))
        self.db.conn.commit(); self.db.event("profile_built", {"cluster": cluster, "version": version, "profile_id": cur.lastrowid})
        return {**profile, "profile_id": cur.lastrowid, "version": version}

    def match(self, cluster: str, resume: str, resume_id: str = "resume-input") -> dict:
        profile_row = self.db.conn.execute("SELECT profile_id,profile_json FROM job_profiles WHERE job_cluster=? ORDER BY version DESC LIMIT 1", (cluster,)).fetchone()
        profile = self.build_profile(cluster) if profile_row is None else json.loads(profile_row[1])
        profile_id = profile.get("profile_id") or profile_row[0]
        resume_skills = {x.name for x in self.extractor.extract(resume)}
        expected = profile["skills"]
        matched = [x for x in expected if x["skill"] in resume_skills]
        missing = [x for x in expected if x["skill"] not in resume_skills]
        denominator = sum(x["weight"] for x in expected) or 1
        score = round(100 * sum(x["weight"] for x in matched) / denominator, 1)
        advice = [f"优先补强：{x['skill']}（岗位画像权重 {x['weight']}）" for x in missing[:3]]
        result = {"job_cluster": cluster, "score": score, "matched_skills": matched, "missing_skills": missing, "advice": advice}
        self.db.conn.execute("INSERT OR REPLACE INTO resumes VALUES(?,?,?)", (resume_id, resume, NOW()))
        self.db.conn.execute("INSERT INTO assessments(resume_id,profile_id,score,matched_json,missing_json,advice_json,created_at) VALUES(?,?,?,?,?,?,?)",
          (resume_id, profile_id, score, json.dumps(matched, ensure_ascii=False), json.dumps(missing, ensure_ascii=False), json.dumps(advice, ensure_ascii=False), NOW()))
        self.db.conn.commit(); self.db.event("resume_matched", {"resume_id": resume_id, "cluster": cluster, "score": score})
        return result
